from fastapi import FastAPI, APIRouter, HTTPException, Header, Body, UploadFile, File
from fastapi.responses import JSONResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import io
import uuid
import logging
import httpx
import openpyxl
from pathlib import Path
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
from datetime import datetime, timezone

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

MONGO_URL = os.environ['MONGO_URL']
DB_NAME = os.environ['DB_NAME']
CLIENT_ID = os.environ['MS_CLIENT_ID']
CLIENT_SECRET = os.environ['MS_CLIENT_SECRET']
TENANT_ID = os.environ.get('MS_TENANT_ID', 'common')
REDIRECT_URI = os.environ['MS_REDIRECT_URI']

SCOPES = "offline_access User.Read Files.Read Files.Read.All"
GRAPH = "https://graph.microsoft.com/v1.0"
AUTHORITY = f"https://login.microsoftonline.com/{TENANT_ID}"

client = AsyncIOMotorClient(MONGO_URL)
db = client[DB_NAME]

app = FastAPI()
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


# ============ Helpers ============

def now_iso():
    return datetime.now(timezone.utc).isoformat()


async def get_session(session_id: str) -> dict:
    if not session_id:
        raise HTTPException(status_code=401, detail="Missing session")
    s = await db.sessions.find_one({"session_id": session_id}, {"_id": 0})
    if not s:
        raise HTTPException(status_code=401, detail="Invalid session")
    return s


async def refresh_token_if_needed(session: dict) -> str:
    """Refresh access token using stored refresh token."""
    async with httpx.AsyncClient() as http:
        resp = await http.post(
            f"{AUTHORITY}/oauth2/v2.0/token",
            data={
                "client_id": CLIENT_ID,
                "client_secret": CLIENT_SECRET,
                "refresh_token": session.get("refresh_token", ""),
                "grant_type": "refresh_token",
                "scope": SCOPES,
            },
            timeout=30,
        )
        data = resp.json()
        if "access_token" not in data:
            raise HTTPException(status_code=401, detail=f"Token refresh failed: {data}")
        await db.sessions.update_one(
            {"session_id": session["session_id"]},
            {"$set": {
                "access_token": data["access_token"],
                "refresh_token": data.get("refresh_token", session.get("refresh_token")),
                "updated_at": now_iso(),
            }},
        )
        return data["access_token"]


async def graph_get(session: dict, url: str, params: dict = None, stream_bytes: bool = False):
    """GET MS Graph with auto-refresh on 401."""
    token = session["access_token"]
    async with httpx.AsyncClient() as http:
        headers = {"Authorization": f"Bearer {token}"}
        resp = await http.get(url, headers=headers, params=params, follow_redirects=True, timeout=60)
        if resp.status_code == 401:
            token = await refresh_token_if_needed(session)
            headers["Authorization"] = f"Bearer {token}"
            resp = await http.get(url, headers=headers, params=params, follow_redirects=True, timeout=60)
        if resp.status_code >= 400:
            raise HTTPException(status_code=resp.status_code, detail=f"Graph error: {resp.text}")
        return resp.content if stream_bytes else resp.json()


# ============ Models ============

class MapCreate(BaseModel):
    name: str
    file_id: str
    file_name: str
    sheet_name: str
    lat_column: str
    lng_column: str
    visible_columns: List[str] = []
    header_row: int = 1  # 1-indexed row that contains headers
    first_col: int = 1   # 1-indexed first column to consider
    status_column: Optional[str] = None
    status_visible_values: List[str] = []
    data_row_from: Optional[int] = None  # 1-based inclusive first data row
    data_row_to: Optional[int] = None    # 1-based inclusive last data row
    source: str = "onedrive"  # "onedrive" or "upload"


class MapUpdate(BaseModel):
    name: Optional[str] = None
    visible_columns: Optional[List[str]] = None
    lat_column: Optional[str] = None
    lng_column: Optional[str] = None
    header_row: Optional[int] = None
    first_col: Optional[int] = None
    status_column: Optional[str] = None
    status_visible_values: Optional[List[str]] = None
    data_row_from: Optional[int] = None
    data_row_to: Optional[int] = None


class PointEdit(BaseModel):
    overrides: Dict[str, Any]


# ============ Auth ============

ALLOWED_REDIRECT_SUFFIXES = (
    ".preview.emergentagent.com",
    ".emergent.host",
    "localhost:3000",
    "127.0.0.1:3000",
)


def _validate_redirect_uri(uri: str) -> bool:
    """Only allow redirect URIs that end at /auth/callback on known hosts."""
    if not uri or not uri.startswith(("http://", "https://")):
        return False
    if not uri.endswith("/auth/callback"):
        return False
    try:
        stripped = uri.split("://", 1)[1]
        host = stripped.split("/", 1)[0]
    except IndexError:
        return False
    return any(host == suf.lstrip(".") or host.endswith(suf) for suf in ALLOWED_REDIRECT_SUFFIXES)


@api_router.get("/auth/microsoft/url")
async def microsoft_auth_url(prompt: Optional[str] = None, redirect_uri: Optional[str] = None):
    # Frontend can pass its own origin so preview & production both work
    if redirect_uri and _validate_redirect_uri(redirect_uri):
        chosen_redirect = redirect_uri
    else:
        chosen_redirect = REDIRECT_URI
    state = str(uuid.uuid4())
    await db.oauth_states.insert_one({
        "state": state,
        "redirect_uri": chosen_redirect,
        "created_at": now_iso(),
    })
    url = (
        f"{AUTHORITY}/oauth2/v2.0/authorize"
        f"?client_id={CLIENT_ID}"
        f"&response_type=code"
        f"&redirect_uri={chosen_redirect}"
        f"&response_mode=query"
        f"&scope={SCOPES}"
        f"&state={state}"
    )
    if prompt in ("login", "select_account", "consent", "none"):
        url += f"&prompt={prompt}"
    return {"url": url, "state": state, "redirect_uri": chosen_redirect}


def _resolve_email(m: dict) -> Optional[str]:
    """Prefer the real mail address over the Azure guest UPN
    (e.g. 'foo_gmail.com#EXT#@tenant.onmicrosoft.com')."""
    mail = (m.get("mail") or "").strip()
    if mail:
        return mail
    others = m.get("otherMails") or []
    if others:
        return others[0]
    upn = (m.get("userPrincipalName") or "").strip()
    if not upn:
        return None
    # Decode guest UPN: 'foo_gmail.com#EXT#@tenant.onmicrosoft.com' -> 'foo@gmail.com'
    if "#EXT#" in upn:
        local = upn.split("#EXT#", 1)[0]
        if "_" in local:
            username, domain = local.rsplit("_", 1)
            return f"{username}@{domain}"
        return local
    return upn


@api_router.post("/auth/microsoft/callback")
async def microsoft_callback(payload: dict = Body(...)):
    code = payload.get("code")
    state = payload.get("state")
    if not code or not state:
        raise HTTPException(status_code=400, detail="Missing code/state")

    state_doc = await db.oauth_states.find_one({"state": state})
    if not state_doc:
        raise HTTPException(status_code=400, detail="Invalid state")
    await db.oauth_states.delete_one({"state": state})

    # Use the same redirect_uri stored during auth URL generation
    used_redirect = state_doc.get("redirect_uri") or REDIRECT_URI

    async with httpx.AsyncClient() as http:
        resp = await http.post(
            f"{AUTHORITY}/oauth2/v2.0/token",
            data={
                "client_id": CLIENT_ID,
                "client_secret": CLIENT_SECRET,
                "code": code,
                "redirect_uri": used_redirect,
                "grant_type": "authorization_code",
                "scope": SCOPES,
            },
            timeout=30,
        )
        data = resp.json()
        if "access_token" not in data:
            raise HTTPException(status_code=400, detail=f"Token exchange failed: {data}")

    # Fetch user profile
    me_resp = None
    # Ask Graph for the actual mail fields too (guest UPNs are ugly)
    async with httpx.AsyncClient() as http:
        me_resp = await http.get(
            f"{GRAPH}/me",
            params={"$select": "id,displayName,mail,userPrincipalName,otherMails"},
            headers={"Authorization": f"Bearer {data['access_token']}"},
            timeout=30,
        )
    me = me_resp.json() if me_resp.status_code == 200 else {}

    session_id = str(uuid.uuid4())
    user_id = me.get("id") or session_id
    session_doc = {
        "session_id": session_id,
        "user_id": user_id,
        "display_name": me.get("displayName"),
        "email": _resolve_email(me),
        "access_token": data["access_token"],
        "refresh_token": data.get("refresh_token"),
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.sessions.insert_one(session_doc)
    return {
        "session_id": session_id,
        "user": {
            "id": user_id,
            "display_name": session_doc["display_name"],
            "email": session_doc["email"],
        },
    }


@api_router.get("/auth/me")
async def auth_me(x_session_id: str = Header(...)):
    s = await get_session(x_session_id)
    return {
        "id": s.get("user_id"),
        "display_name": s.get("display_name"),
        "email": s.get("email"),
    }


@api_router.post("/auth/logout")
async def auth_logout(x_session_id: str = Header(...)):
    await db.sessions.delete_one({"session_id": x_session_id})
    return {"ok": True}


# ============ Local Auth (email + password) ============

import bcrypt
import re


class LocalSignup(BaseModel):
    email: str
    password: str
    display_name: Optional[str] = None


class LocalLogin(BaseModel):
    email: str
    password: str


def _hash_password(pw: str) -> str:
    return bcrypt.hashpw(pw.encode("utf-8"), bcrypt.gensalt()).decode("utf-8")


def _verify_password(pw: str, hashed: str) -> bool:
    try:
        return bcrypt.checkpw(pw.encode("utf-8"), hashed.encode("utf-8"))
    except (ValueError, TypeError):
        return False


EMAIL_RE = re.compile(r"^[A-Za-z0-9._%+\-]+@[A-Za-z0-9.\-]+\.[A-Za-z]{2,}$")


async def _create_local_session(user_doc: dict) -> str:
    session_id = str(uuid.uuid4())
    await db.sessions.insert_one({
        "session_id": session_id,
        "user_id": user_doc["id"],
        "auth_type": "local",
        "display_name": user_doc.get("display_name"),
        "email": user_doc["email"],
        "access_token": None,
        "refresh_token": None,
        "created_at": now_iso(),
        "updated_at": now_iso(),
    })
    return session_id


@api_router.post("/auth/local/signup")
async def local_signup(payload: LocalSignup):
    email = (payload.email or "").strip().lower()
    if not EMAIL_RE.match(email):
        raise HTTPException(status_code=400, detail="Email inválido")
    if len(payload.password) < 8:
        raise HTTPException(status_code=400, detail="La contraseña debe tener al menos 8 caracteres")
    existing = await db.users.find_one({"email": email})
    if existing:
        raise HTTPException(status_code=409, detail="Ya existe una cuenta con ese email")
    user_id = f"local:{uuid.uuid4()}"
    user_doc = {
        "id": user_id,
        "email": email,
        "password_hash": _hash_password(payload.password),
        "display_name": (payload.display_name or email.split("@")[0]).strip() or email.split("@")[0],
        "created_at": now_iso(),
    }
    await db.users.insert_one(user_doc)
    session_id = await _create_local_session(user_doc)
    return {
        "session_id": session_id,
        "user": {"id": user_id, "email": email, "display_name": user_doc["display_name"]},
    }


@api_router.post("/auth/local/login")
async def local_login(payload: LocalLogin):
    email = (payload.email or "").strip().lower()
    user = await db.users.find_one({"email": email})
    # Generic message to prevent user enumeration
    if not user or not _verify_password(payload.password, user.get("password_hash", "")):
        raise HTTPException(status_code=401, detail="Email o contraseña incorrectos")
    session_id = await _create_local_session(user)
    return {
        "session_id": session_id,
        "user": {"id": user["id"], "email": user["email"], "display_name": user.get("display_name")},
    }


# ============ Batch update maps source ============

class BatchRefreshPayload(BaseModel):
    map_ids: List[str]
    upload_id: str


@api_router.post("/maps/batch-refresh-source")
async def batch_refresh_source(payload: BatchRefreshPayload, x_session_id: str = Header(...)):
    """Update multiple maps to point to a new uploaded file. The user must own each map
    AND own the new upload. This is how a user re-uploads a fresh version of the .xlsx
    and applies it to all maps that used the previous version."""
    s = await get_session(x_session_id)
    upload = await db.uploads.find_one({"id": payload.upload_id, "user_id": s["user_id"]}, {"data": 0})
    if not upload:
        raise HTTPException(status_code=404, detail="Archivo subido no encontrado")
    if not payload.map_ids:
        return {"updated": 0, "map_ids": []}
    # Verify ownership of all maps
    owned = await db.maps.find(
        {"id": {"$in": payload.map_ids}, "user_id": s["user_id"]},
        {"id": 1, "_id": 0},
    ).to_list(500)
    owned_ids = [m["id"] for m in owned]
    if not owned_ids:
        raise HTTPException(status_code=404, detail="Ninguno de los mapas es tuyo")
    await db.maps.update_many(
        {"id": {"$in": owned_ids}, "user_id": s["user_id"]},
        {"$set": {
            "source": "upload",
            "file_id": payload.upload_id,
            "file_name": upload["filename"],
            "updated_at": now_iso(),
        }},
    )
    return {"updated": len(owned_ids), "map_ids": owned_ids}


# ============ OneDrive Excel ============

@api_router.get("/onedrive/files")
async def list_excel_files(x_session_id: str = Header(...)):
    s = await get_session(x_session_id)
    seen = {}

    # 1) Root children (top-level files in OneDrive root)
    try:
        data = await graph_get(s, f"{GRAPH}/me/drive/root/children", params={"$top": "200"})
        for f in data.get("value", []):
            name = f.get("name", "")
            if name.lower().endswith(".xlsx") and "file" in f:
                seen[f["id"]] = f
    except HTTPException as e:
        logger.warning(f"root/children failed: {e.detail}")

    # 2) Full-drive search (indexed, includes subfolders)
    try:
        data = await graph_get(s, f"{GRAPH}/me/drive/root/search(q='.xlsx')")
        for f in data.get("value", []):
            name = f.get("name", "")
            if name.lower().endswith(".xlsx") and "file" in f:
                seen[f["id"]] = f
    except HTTPException as e:
        logger.warning(f"search failed: {e.detail}")

    # 3) Recent files (catches files opened recently, even shared)
    try:
        data = await graph_get(s, f"{GRAPH}/me/drive/recent", params={"$top": "50"})
        for f in data.get("value", []):
            name = f.get("name", "")
            if name.lower().endswith(".xlsx") and "file" in f:
                seen.setdefault(f["id"], f)
    except HTTPException as e:
        logger.warning(f"recent failed: {e.detail}")

    files = [
        {
            "id": f["id"],
            "name": f.get("name", ""),
            "size": f.get("size"),
            "modified": f.get("lastModifiedDateTime"),
            "web_url": f.get("webUrl"),
            "path": (f.get("parentReference") or {}).get("path"),
        }
        for f in seen.values()
    ]
    files.sort(key=lambda x: x.get("modified") or "", reverse=True)
    return {"files": files, "count": len(files)}


async def _load_workbook(session: dict, item_id: str) -> openpyxl.Workbook:
    content = await graph_get(session, f"{GRAPH}/me/drive/items/{item_id}/content", stream_bytes=True)
    wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    return wb


async def _load_uploaded_workbook(upload_id: str) -> openpyxl.Workbook:
    """Load an .xlsx that was uploaded from device (stored in Mongo)."""
    doc = await db.uploads.find_one({"id": upload_id}, {"data": 1})
    if not doc or not doc.get("data"):
        raise HTTPException(status_code=404, detail="Archivo subido no encontrado")
    return openpyxl.load_workbook(io.BytesIO(doc["data"]), read_only=True, data_only=True)


async def _load_workbook_for_map(m: dict, session: dict) -> openpyxl.Workbook:
    """Dispatch workbook loading based on the map's source."""
    if m.get("source") == "upload":
        return await _load_uploaded_workbook(m["file_id"])
    return await _load_workbook(session, m["file_id"])


# ============ Uploads (files from user's device) ============

MAX_UPLOAD_SIZE = 15 * 1024 * 1024  # 15 MB — BSON limit is 16MB


@api_router.post("/uploads/excel")
async def upload_excel(file: UploadFile = File(...), x_session_id: str = Header(...)):
    s = await get_session(x_session_id)
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Solo se aceptan archivos .xlsx")
    data = await file.read()
    if len(data) == 0:
        raise HTTPException(status_code=400, detail="El archivo está vacío")
    if len(data) > MAX_UPLOAD_SIZE:
        raise HTTPException(status_code=413, detail=f"El archivo excede el límite de {MAX_UPLOAD_SIZE // (1024*1024)} MB")
    # Validate it's a proper xlsx
    try:
        openpyxl.load_workbook(io.BytesIO(data), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Archivo Excel inválido: {e}")

    upload_id = str(uuid.uuid4())
    await db.uploads.insert_one({
        "id": upload_id,
        "user_id": s["user_id"],
        "filename": file.filename,
        "content_type": file.content_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "size": len(data),
        "data": data,
        "created_at": now_iso(),
    })
    return {"id": upload_id, "filename": file.filename, "size": len(data)}


@api_router.get("/uploads/files")
async def list_uploads(x_session_id: str = Header(...)):
    s = await get_session(x_session_id)
    items = await db.uploads.find(
        {"user_id": s["user_id"]},
        {"_id": 0, "data": 0},
    ).sort("created_at", -1).to_list(200)
    return {"files": items}


@api_router.delete("/uploads/files/{upload_id}")
async def delete_upload(upload_id: str, x_session_id: str = Header(...)):
    s = await get_session(x_session_id)
    res = await db.uploads.delete_one({"id": upload_id, "user_id": s["user_id"]})
    if res.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Upload not found")
    return {"ok": True}


@api_router.get("/uploads/files/{upload_id}/sheets")
async def upload_sheets(upload_id: str, x_session_id: str = Header(...)):
    await get_session(x_session_id)
    wb = await _load_uploaded_workbook(upload_id)
    return {"sheets": wb.sheetnames}


@api_router.get("/uploads/files/{upload_id}/sheets/{sheet_name}/preview")
async def upload_sheet_preview(
    upload_id: str,
    sheet_name: str,
    max_rows: int = 25,
    max_cols: int = 20,
    x_session_id: str = Header(...),
):
    await get_session(x_session_id)
    wb = await _load_uploaded_workbook(upload_id)
    return _grid_preview(wb, sheet_name, max_rows, max_cols)


@api_router.get("/uploads/files/{upload_id}/sheets/{sheet_name}/data")
async def upload_sheet_data(
    upload_id: str,
    sheet_name: str,
    header_row: int = 1,
    first_col: int = 1,
    x_session_id: str = Header(...),
):
    await get_session(x_session_id)
    wb = await _load_uploaded_workbook(upload_id)
    return _read_sheet_data(wb, sheet_name, header_row, first_col)


@api_router.get("/onedrive/files/{item_id}/sheets")
async def list_sheets(item_id: str, x_session_id: str = Header(...)):
    s = await get_session(x_session_id)
    wb = await _load_workbook(s, item_id)
    return {"sheets": wb.sheetnames}


def _parse_sheet(
    ws,
    header_row: int = 1,
    first_col: int = 1,
):
    """Read sheet with configurable start row/col. Returns (headers, data_rows).
    header_row is 1-indexed; first_col is 1-indexed.
    """
    header_row = max(1, header_row)
    first_col = max(1, first_col)
    all_rows = list(ws.iter_rows(values_only=True))
    if len(all_rows) < header_row:
        return [], []
    header_line = all_rows[header_row - 1]
    # Slice from first_col-1 and drop trailing None headers
    header_slice = list(header_line)[first_col - 1:]
    # Trim trailing empty header cells
    while header_slice and (header_slice[-1] is None or str(header_slice[-1]).strip() == ""):
        header_slice.pop()
    headers = [str(h) if h is not None else f"col_{i}" for i, h in enumerate(header_slice)]

    data_rows = []
    for r in all_rows[header_row:]:
        row_slice = list(r)[first_col - 1: first_col - 1 + len(headers)]
        vals = row_slice + [None] * (len(headers) - len(row_slice))
        # Skip rows that are entirely empty
        if all(v is None or (isinstance(v, str) and v.strip() == "") for v in vals):
            continue
        data_rows.append(vals)
    return headers, data_rows


@api_router.get("/onedrive/files/{item_id}/sheets/{sheet_name}/preview")
async def sheet_preview(
    item_id: str,
    sheet_name: str,
    max_rows: int = 25,
    max_cols: int = 20,
    x_session_id: str = Header(...),
):
    """Return raw grid (no header assumption) so frontend can render an Excel-like preview
    where the user clicks a cell to define the data start (header_row, first_col)."""
    s = await get_session(x_session_id)
    wb = await _load_workbook(s, item_id)
    return _grid_preview(wb, sheet_name, max_rows, max_cols)


def _grid_preview(wb, sheet_name: str, max_rows: int, max_cols: int):
    if sheet_name not in wb.sheetnames:
        raise HTTPException(status_code=404, detail="Sheet not found")
    ws = wb[sheet_name]
    grid = []
    for row in ws.iter_rows(values_only=True):
        row_cells = []
        for v in list(row)[:max_cols]:
            if isinstance(v, datetime):
                v = v.isoformat()
            row_cells.append(v)
        while len(row_cells) < max_cols:
            row_cells.append(None)
        grid.append(row_cells)
        if len(grid) >= max_rows:
            break

    suggested_header_row = 1
    suggested_first_col = 1
    for r_idx, row in enumerate(grid):
        non_empty_indices = [
            c_idx for c_idx, v in enumerate(row)
            if v is not None and str(v).strip() != ""
        ]
        if len(non_empty_indices) >= 3:
            suggested_header_row = r_idx + 1
            suggested_first_col = non_empty_indices[0] + 1
            break

    return {
        "grid": grid,
        "rows": len(grid),
        "cols": max_cols,
        "suggested_header_row": suggested_header_row,
        "suggested_first_col": suggested_first_col,
    }


@api_router.get("/onedrive/files/{item_id}/sheets/{sheet_name}/data")
async def sheet_data(
    item_id: str,
    sheet_name: str,
    header_row: int = 1,
    first_col: int = 1,
    x_session_id: str = Header(...),
):
    s = await get_session(x_session_id)
    wb = await _load_workbook(s, item_id)
    return _read_sheet_data(wb, sheet_name, header_row, first_col)


def _read_sheet_data(wb, sheet_name: str, header_row: int, first_col: int):
    if sheet_name not in wb.sheetnames:
        raise HTTPException(status_code=404, detail="Sheet not found")
    ws = wb[sheet_name]
    headers, data_rows = _parse_sheet(ws, header_row=header_row, first_col=first_col)

    if not headers:
        raise HTTPException(status_code=400, detail="La hoja no tiene columnas en el rango indicado.")

    sample_rows = []
    for vals in data_rows[:3]:
        d = {}
        for h, v in zip(headers, vals):
            if isinstance(v, datetime):
                v = v.isoformat()
            d[h] = v
        sample_rows.append(d)

    lat_candidates = [h for h in headers if any(k in h.lower() for k in ["lat", "latitud"])]
    lng_candidates = [h for h in headers if any(k in h.lower() for k in ["lon", "lng", "longitud"])]
    suggested_lat = lat_candidates[0] if lat_candidates else (headers[-2] if len(headers) >= 2 else None)
    suggested_lng = lng_candidates[0] if lng_candidates else (headers[-1] if len(headers) >= 1 else None)

    return {
        "headers": headers,
        "sample_rows": sample_rows,
        "suggested_lat_column": suggested_lat,
        "suggested_lng_column": suggested_lng,
        "row_count": len(data_rows),
    }


# ============ Maps CRUD ============

def _norm_email(e: Optional[str]) -> Optional[str]:
    return (e or "").strip().lower() or None


async def _get_editable_map(map_id: str, session: dict) -> dict:
    """Return the map if session user is owner OR editor. Raises 404 otherwise."""
    user_id = session["user_id"]
    email = _norm_email(session.get("email"))
    query = {"id": map_id, "$or": [{"user_id": user_id}]}
    if email:
        query["$or"].append({"editor_emails": email})
    m = await db.maps.find_one(query)
    if not m:
        raise HTTPException(status_code=404, detail="Map not found")
    return m


async def _get_owned_map(map_id: str, session: dict) -> dict:
    """Return the map only if session user is the owner. Raises 404 otherwise."""
    m = await db.maps.find_one({"id": map_id, "user_id": session["user_id"]})
    if not m:
        raise HTTPException(status_code=404, detail="Map not found or not owned by you")
    return m


def _decorate_map(m: dict, session: dict) -> dict:
    """Add computed fields (is_owner) to map response."""
    m.pop("_id", None)
    m["is_owner"] = m.get("user_id") == session["user_id"]
    return m


@api_router.post("/maps")
async def create_map(payload: MapCreate, x_session_id: str = Header(...)):
    s = await get_session(x_session_id)
    map_id = str(uuid.uuid4())
    owner_email = _norm_email(s.get("email"))
    doc = {
        "id": map_id,
        "user_id": s["user_id"],
        "owner_email": owner_email,
        "owner_display_name": s.get("display_name"),
        "editor_emails": [],
        "name": payload.name,
        "source": payload.source or "onedrive",
        "file_id": payload.file_id,
        "file_name": payload.file_name,
        "sheet_name": payload.sheet_name,
        "lat_column": payload.lat_column,
        "lng_column": payload.lng_column,
        "visible_columns": payload.visible_columns,
        "header_row": payload.header_row,
        "first_col": payload.first_col,
        "status_column": payload.status_column,
        "status_visible_values": payload.status_visible_values,
        "point_overrides": {},
        "data_row_from": payload.data_row_from,
        "data_row_to": payload.data_row_to,
        "is_public": False,
        "share_token": None,
        "cached_rows": [],
        "cached_at": None,
        "cached_headers": [],
        "created_at": now_iso(),
        "updated_at": now_iso(),
    }
    await db.maps.insert_one(doc)
    return _decorate_map(doc, s)


@api_router.get("/maps")
async def list_maps(x_session_id: str = Header(...)):
    s = await get_session(x_session_id)
    email = _norm_email(s.get("email"))
    query = {"$or": [{"user_id": s["user_id"]}]}
    if email:
        query["$or"].append({"editor_emails": email})
    items = await db.maps.find(query, {"_id": 0}).sort("created_at", -1).to_list(500)
    for m in items:
        m["is_owner"] = m.get("user_id") == s["user_id"]
    return {"maps": items}


@api_router.get("/maps/{map_id}")
async def get_map(map_id: str, x_session_id: str = Header(...)):
    s = await get_session(x_session_id)
    m = await _get_editable_map(map_id, s)
    return _decorate_map(m, s)


@api_router.patch("/maps/{map_id}")
async def update_map(map_id: str, payload: MapUpdate, x_session_id: str = Header(...)):
    s = await get_session(x_session_id)
    await _get_editable_map(map_id, s)
    # Use exclude_unset=True so clients can explicitly reset fields to null
    updates = payload.model_dump(exclude_unset=True)
    if not updates:
        raise HTTPException(status_code=400, detail="No updates")
    updates["updated_at"] = now_iso()
    await db.maps.update_one({"id": map_id}, {"$set": updates})
    m = await db.maps.find_one({"id": map_id}, {"_id": 0})
    return _decorate_map(m, s)


@api_router.delete("/maps/{map_id}")
async def delete_map(map_id: str, x_session_id: str = Header(...)):
    s = await get_session(x_session_id)
    await _get_owned_map(map_id, s)
    await db.maps.delete_one({"id": map_id})
    return {"ok": True}


@api_router.put("/maps/{map_id}/points/{row_index}")
async def update_point(map_id: str, row_index: int, payload: PointEdit, x_session_id: str = Header(...)):
    s = await get_session(x_session_id)
    await _get_editable_map(map_id, s)
    key = f"point_overrides.{row_index}"
    await db.maps.update_one(
        {"id": map_id},
        {"$set": {key: payload.overrides, "updated_at": now_iso()}},
    )
    updated = await db.maps.find_one({"id": map_id}, {"_id": 0})
    return _decorate_map(updated, s)


@api_router.delete("/maps/{map_id}/points/{row_index}")
async def reset_point(map_id: str, row_index: int, x_session_id: str = Header(...)):
    s = await get_session(x_session_id)
    await _get_editable_map(map_id, s)
    key = f"point_overrides.{row_index}"
    await db.maps.update_one({"id": map_id}, {"$unset": {key: ""}, "$set": {"updated_at": now_iso()}})
    updated = await db.maps.find_one({"id": map_id}, {"_id": 0})
    return _decorate_map(updated, s)


def _build_map_rows(headers, data_rows, m):
    """Given headers + raw data_rows (list of lists) + map doc, build normalized rows."""
    lat_col = m.get("lat_column")
    lng_col = m.get("lng_column")
    status_col = m.get("status_column")
    status_visible = set(m.get("status_visible_values") or [])
    row_from = m.get("data_row_from")
    row_to = m.get("data_row_to")
    # Backwards-compat fallback
    if (not lat_col or lat_col not in headers) and len(headers) >= 2:
        lat_col = headers[-2]
    if (not lng_col or lng_col not in headers) and len(headers) >= 1:
        lng_col = headers[-1]
    if lat_col not in headers or lng_col not in headers:
        raise HTTPException(
            status_code=400,
            detail=f"Columnas de coordenadas ({lat_col}, {lng_col}) no encontradas. Disponibles: {headers}",
        )
    lat_idx = headers.index(lat_col)
    lng_idx = headers.index(lng_col)
    overrides = m.get("point_overrides", {}) or {}
    rows = []
    for idx, vals in enumerate(data_rows):
        # Row range filter (1-based, inclusive)
        row_num = idx + 1
        if row_from is not None and row_num < row_from:
            continue
        if row_to is not None and row_num > row_to:
            continue
        try:
            lat_raw = vals[lat_idx]
            lat = float(lat_raw) if lat_raw is not None and str(lat_raw).strip() != "" else None
        except (ValueError, TypeError):
            lat = None
        try:
            lng_raw = vals[lng_idx]
            lng = float(lng_raw) if lng_raw is not None and str(lng_raw).strip() != "" else None
        except (ValueError, TypeError):
            lng = None
        data_dict = {}
        for h, v in zip(headers, vals):
            if isinstance(v, datetime):
                v = v.isoformat()
            data_dict[h] = v
        ov = overrides.get(str(idx)) or {}
        for k, v in ov.items():
            if k in (lat_col, lng_col):
                continue
            data_dict[k] = v
        # Status-based visibility
        is_visible = True
        if status_col and status_col in headers:
            status_val = data_dict.get(status_col)
            is_empty = status_val is None or (isinstance(status_val, str) and status_val.strip() == "")
            status_str = "" if status_val is None else str(status_val)
            # If no status_visible values configured, everything is visible
            if status_visible:
                if is_empty:
                    is_visible = "__EMPTY__" in status_visible
                else:
                    is_visible = status_str in status_visible
        rows.append({
            "row_index": idx,
            "lat": lat,
            "lng": lng,
            "data": data_dict,
            "edited": bool(ov),
            "visible": is_visible,
        })
    return rows, lat_col, lng_col


@api_router.get("/maps/{map_id}/data")
async def map_data(map_id: str, x_session_id: str = Header(...)):
    """Combine map config with live Excel data + overrides applied."""
    s = await get_session(x_session_id)
    m = await _get_editable_map(map_id, s)
    m.pop("_id", None)
    # Use owner's session to fetch the Excel — editors don't have access to owner's OneDrive.
    if m.get("source") == "upload":
        # Uploaded files don't need MS Graph at all
        graph_session = s
    elif m.get("user_id") != s["user_id"]:
        owner_session = await db.sessions.find_one(
            {"user_id": m.get("user_id")}, sort=[("updated_at", -1)]
        )
        if not owner_session:
            raise HTTPException(
                status_code=409,
                detail="El dueño del mapa debe iniciar sesión al menos una vez para que puedas ver los datos actualizados.",
            )
        graph_session = owner_session
    else:
        graph_session = s
    wb = await _load_workbook_for_map(m, graph_session)
    if m["sheet_name"] not in wb.sheetnames:
        raise HTTPException(status_code=404, detail="Sheet no longer exists in workbook")
    ws = wb[m["sheet_name"]]
    header_row = m.get("header_row") or 1
    first_col = m.get("first_col") or 1
    headers, data_rows = _parse_sheet(ws, header_row=header_row, first_col=first_col)

    rows, lat_col, lng_col = _build_map_rows(headers, data_rows, m)

    # Cache snapshot for public views
    await db.maps.update_one(
        {"id": map_id},
        {"$set": {
            "cached_rows": [{"row_index": r["row_index"], "lat": r["lat"], "lng": r["lng"],
                             "data": r["data"], "edited": r["edited"], "visible": r["visible"]}
                            for r in rows],
            "cached_headers": headers,
            "cached_at": now_iso(),
        }},
    )

    # Get unique status values if status_column configured
    status_values = []
    has_empty = False
    status_col = m.get("status_column")
    if status_col and status_col in headers:
        seen_vals = set()
        for r in rows:
            v = r["data"].get(status_col)
            if v is None or (isinstance(v, str) and v.strip() == ""):
                has_empty = True
                continue
            s_val = str(v)
            if s_val not in seen_vals:
                seen_vals.add(s_val)
                status_values.append(s_val)

    return {
        "map": _decorate_map(m, s),
        "headers": headers,
        "lat_column": lat_col,
        "lng_column": lng_col,
        "status_column": status_col,
        "status_values": status_values,
        "status_has_empty": has_empty,
        "rows": rows,
    }


# ============ Editors (collaboration) ============

class EditorAdd(BaseModel):
    email: str


@api_router.get("/maps/{map_id}/editors")
async def list_editors(map_id: str, x_session_id: str = Header(...)):
    s = await get_session(x_session_id)
    m = await _get_editable_map(map_id, s)
    return {
        "owner_email": m.get("owner_email"),
        "owner_display_name": m.get("owner_display_name"),
        "editors": m.get("editor_emails") or [],
        "is_owner": m.get("user_id") == s["user_id"],
    }


@api_router.post("/maps/{map_id}/editors")
async def add_editor(map_id: str, payload: EditorAdd, x_session_id: str = Header(...)):
    s = await get_session(x_session_id)
    m = await _get_owned_map(map_id, s)
    email = _norm_email(payload.email)
    if not email or "@" not in email:
        raise HTTPException(status_code=400, detail="Email inválido")
    if email == _norm_email(m.get("owner_email")):
        raise HTTPException(status_code=400, detail="El dueño ya tiene acceso")
    current = set(m.get("editor_emails") or [])
    current.add(email)
    await db.maps.update_one(
        {"id": map_id},
        {"$set": {"editor_emails": sorted(current), "updated_at": now_iso()}},
    )
    return {"editors": sorted(current)}


@api_router.delete("/maps/{map_id}/editors/{email}")
async def remove_editor(map_id: str, email: str, x_session_id: str = Header(...)):
    s = await get_session(x_session_id)
    m = await _get_owned_map(map_id, s)
    normalized = _norm_email(email)
    current = [e for e in (m.get("editor_emails") or []) if e != normalized]
    await db.maps.update_one(
        {"id": map_id},
        {"$set": {"editor_emails": current, "updated_at": now_iso()}},
    )
    return {"editors": current}


# ============ Share / Public ============

@api_router.post("/maps/{map_id}/share")
async def enable_share(map_id: str, x_session_id: str = Header(...)):
    s = await get_session(x_session_id)
    m = await _get_owned_map(map_id, s)
    token = m.get("share_token") or uuid.uuid4().hex
    await db.maps.update_one(
        {"id": map_id},
        {"$set": {"is_public": True, "share_token": token, "updated_at": now_iso()}},
    )
    updated = await db.maps.find_one({"id": map_id}, {"_id": 0})
    return {"share_token": token, "is_public": True, "map": _decorate_map(updated, s)}


@api_router.post("/maps/{map_id}/share/rotate")
async def rotate_share(map_id: str, x_session_id: str = Header(...)):
    s = await get_session(x_session_id)
    await _get_owned_map(map_id, s)
    token = uuid.uuid4().hex
    await db.maps.update_one(
        {"id": map_id},
        {"$set": {"is_public": True, "share_token": token, "updated_at": now_iso()}},
    )
    return {"share_token": token, "is_public": True}


@api_router.delete("/maps/{map_id}/share")
async def disable_share(map_id: str, x_session_id: str = Header(...)):
    s = await get_session(x_session_id)
    await _get_owned_map(map_id, s)
    await db.maps.update_one(
        {"id": map_id},
        {"$set": {"is_public": False, "updated_at": now_iso()}},
    )
    return {"is_public": False}


VISIT_TTL_HOURS = 12


def _active_visits(m: dict) -> Dict[str, str]:
    """Return only visits whose timestamp is within the TTL window."""
    visits = m.get("public_visits") or {}
    if not visits:
        return {}
    now = datetime.now(timezone.utc)
    active = {}
    for row_idx, iso_ts in visits.items():
        try:
            ts = datetime.fromisoformat(iso_ts.replace("Z", "+00:00"))
            if ts.tzinfo is None:
                ts = ts.replace(tzinfo=timezone.utc)
            if (now - ts).total_seconds() < VISIT_TTL_HOURS * 3600:
                active[str(row_idx)] = iso_ts
        except (ValueError, AttributeError):
            continue
    return active


@api_router.get("/public/maps/{share_token}")
async def public_map(share_token: str):
    """Serves cached snapshot of a shared map. No auth required."""
    m = await db.maps.find_one(
        {"share_token": share_token, "is_public": True},
        {"_id": 0, "user_id": 0, "file_id": 0, "point_overrides": 0},
    )
    if not m:
        raise HTTPException(status_code=404, detail="Mapa no encontrado o no compartido")
    cached_rows = m.get("cached_rows") or []
    cached_headers = m.get("cached_headers") or []
    active_visits = _active_visits(m)
    # Attach 'visited' flag to each row
    rows_with_visits = []
    for r in cached_rows:
        row_copy = dict(r)
        row_copy["visited"] = str(r.get("row_index")) in active_visits
        row_copy["visited_at"] = active_visits.get(str(r.get("row_index")))
        rows_with_visits.append(row_copy)
    return {
        "map": {
            "id": m.get("id"),
            "name": m.get("name"),
            "file_name": m.get("file_name"),
            "sheet_name": m.get("sheet_name"),
            "visible_columns": m.get("visible_columns", []),
            "status_column": m.get("status_column"),
            "status_visible_values": m.get("status_visible_values", []),
        },
        "headers": cached_headers,
        "lat_column": m.get("lat_column"),
        "lng_column": m.get("lng_column"),
        "status_column": m.get("status_column"),
        "rows": rows_with_visits,
        "cached_at": m.get("cached_at"),
        "visit_ttl_hours": VISIT_TTL_HOURS,
    }


@api_router.post("/public/maps/{share_token}/visits/{row_index}")
async def public_mark_visit(share_token: str, row_index: int):
    """Anonymous visitor marks a point as visited. Auto-expires after VISIT_TTL_HOURS."""
    m = await db.maps.find_one(
        {"share_token": share_token, "is_public": True},
        {"id": 1, "public_visits": 1},
    )
    if not m:
        raise HTTPException(status_code=404, detail="Mapa no encontrado o no compartido")
    now = now_iso()
    # Clean expired entries and add the new one
    active = _active_visits(m)
    active[str(row_index)] = now
    await db.maps.update_one(
        {"id": m["id"]},
        {"$set": {"public_visits": active}},
    )
    return {
        "row_index": row_index,
        "visited": True,
        "visited_at": now,
        "expires_in_hours": VISIT_TTL_HOURS,
    }


@api_router.delete("/public/maps/{share_token}/visits/{row_index}")
async def public_clear_visit(share_token: str, row_index: int):
    """Visitor unchecks visited state."""
    m = await db.maps.find_one(
        {"share_token": share_token, "is_public": True},
        {"id": 1, "public_visits": 1},
    )
    if not m:
        raise HTTPException(status_code=404, detail="Mapa no encontrado o no compartido")
    active = _active_visits(m)
    active.pop(str(row_index), None)
    await db.maps.update_one(
        {"id": m["id"]},
        {"$set": {"public_visits": active}},
    )
    return {"row_index": row_index, "visited": False}


@api_router.get("/")
async def root():
    return {"message": "Excel Maps API"}


app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.on_event("startup")
async def startup_indexes():
    try:
        await db.users.create_index("email", unique=True)
    except Exception as e:
        logger.warning(f"users email index: {e}")


@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
